# %%
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pyproj import Transformer

# ====== パラメータ（ここだけ編集すればOK） ======

# Trajectory Excel（添付のやつ）
WELL_TRAJ_XLSX = Path(
	'/home/dcuser/daseventnet/data/silixa/FORGE_Well_Trajectories_GES_April2022.xlsx'
)
SHEET_A = 'Trajectory78A-32'
SHEET_B = 'Trajectory78B-32'

# 座標系（FORGEならだいたい NAD83 / UTM zone 12N）
INPUT_EPSG = 26912

# Tap-testで決まった「井口(深度0)」と「深部端(深度max)」のキャリブレーション（単位: m）
W78A = dict(
	well='78A-32',
	E_m=335780.84,
	N_m=4262991.99,
	elev_head_m=1701.92,
	depth_bottom_m=989.90,
	ch_shallow=1062,
	ch_deep=92,
)

W78B = dict(
	well='78B-32',
	E_m=335865.45,
	N_m=4262983.53,
	elev_head_m=1705.62,
	depth_bottom_m=1193.42,
	ch_shallow=1216,
	ch_deep=2385,
)

# ---- Report-correct well ranges (0-based, inclusive) ----
APPLY_WELL_AB_KEEP = True
WELL_A_KEEP_0BASED_INCL = (92, 1062)
WELL_B_KEEP_0BASED_INCL = (1216, 2385)

# ローカル座標の原点（おすすめ: 78Bの井口）
ORIGIN_E_m = W78B['E_m']
ORIGIN_N_m = W78B['N_m']

# 出力
OUT_DIR = Path('/workspace/data/station/forge')
OUT_CSV = OUT_DIR / 'forge_das_station_metadata.csv'
OUT_MEMO = OUT_DIR / 'forge_das_station_metadata_README.md'

FT_TO_M = 0.3048


def _find_cols(header: list[str], keys: list[str]) -> int:
	for i, h in enumerate(header):
		hl = h.strip().lower()
		ok = True
		for k in keys:
			if k not in hl:
				ok = False
				break
		if ok:
			return i
	raise ValueError(f'column not found for keys={keys}. header={header}')


def read_trajectory_offsets_ft(
	xlsx: Path, sheet: str
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
	if not xlsx.exists():
		raise FileNotFoundError(f'Trajectory xlsx not found: {xlsx}')

	wb = load_workbook(str(xlsx), read_only=True, data_only=True)
	if sheet not in wb.sheetnames:
		raise ValueError(f'sheet not found: {sheet}. sheets={wb.sheetnames}')

	ws = wb[sheet]

	header_row = None
	header_vals = None
	for r in range(1, 40):
		row = [ws.cell(row=r, column=c).value for c in range(1, 40)]
		row_s = [str(x).strip() if x is not None else '' for x in row]
		row_l = [s.lower() for s in row_s]
		if (
			any('e-w' in s for s in row_l)
			and any('n-s' in s for s in row_l)
			and any('tvd' in s for s in row_l)
		):
			header_row = r
			header_vals = row_s
			break

	if header_row is None or header_vals is None:
		raise ValueError(f'trajectory header not found in sheet={sheet}')

	col_ew = _find_cols(header_vals, ['e-w'])
	col_ns = _find_cols(header_vals, ['n-s'])
	col_tvd = _find_cols(header_vals, ['tvd'])

	ew_list: list[float] = []
	ns_list: list[float] = []
	tvd_list: list[float] = []

	for r in range(header_row + 1, header_row + 2000):
		v_ew = ws.cell(row=r, column=col_ew + 1).value
		v_ns = ws.cell(row=r, column=col_ns + 1).value
		v_tvd = ws.cell(row=r, column=col_tvd + 1).value

		if v_ew is None and v_ns is None and v_tvd is None:
			if len(tvd_list) > 10:
				break
			continue

		if (
			not isinstance(v_ew, (int, float))
			or not isinstance(v_ns, (int, float))
			or not isinstance(v_tvd, (int, float))
		):
			continue

		ew_list.append(float(v_ew))
		ns_list.append(float(v_ns))
		tvd_list.append(float(v_tvd))

	if len(tvd_list) < 3:
		raise ValueError(f'trajectory rows too few in sheet={sheet}')

	ew = np.asarray(ew_list, dtype=float)
	ns = np.asarray(ns_list, dtype=float)
	tvd = np.asarray(tvd_list, dtype=float)

	order = np.argsort(tvd)
	ew = ew[order]
	ns = ns[order]
	tvd = tvd[order]

	return ew, ns, tvd


def channel_to_depth_m(ch: np.ndarray, cfg: dict) -> np.ndarray:
	ch0 = int(cfg['ch_shallow'])
	ch1 = int(cfg['ch_deep'])
	depth1 = float(cfg['depth_bottom_m'])
	if ch1 == ch0:
		raise ValueError(f'bad calibration: ch_shallow==ch_deep for well={cfg["well"]}')
	return (ch.astype(float) - float(ch0)) * (depth1 / float(ch1 - ch0))


def depth_to_offsets_m(
	depth_m: np.ndarray, ew_ft: np.ndarray, ns_ft: np.ndarray, tvd_ft: np.ndarray
) -> tuple[np.ndarray, np.ndarray]:
	tvd_m = tvd_ft.astype(float) * FT_TO_M
	ew_m = ew_ft.astype(float) * FT_TO_M
	ns_m = ns_ft.astype(float) * FT_TO_M

	dmin = float(tvd_m.min())
	dmax = float(tvd_m.max())
	if float(depth_m.min()) < dmin - 1.0 or float(depth_m.max()) > dmax + 1.0:
		raise ValueError(
			f'depth out of trajectory bounds: depth=[{depth_m.min():.2f},{depth_m.max():.2f}] tvd=[{dmin:.2f},{dmax:.2f}]'
		)

	offE = np.interp(depth_m.astype(float), tvd_m, ew_m)
	offN = np.interp(depth_m.astype(float), tvd_m, ns_m)
	return offE.astype(float), offN.astype(float)


def make_well_df(
	cfg: dict,
	origin_E_m: float,
	origin_N_m: float,
	*,
	keep_0based_incl: tuple[int, int] | None,
	index_offset: int,
	station_prefix: str,
	traj_ew_ft: np.ndarray,
	traj_ns_ft: np.ndarray,
	traj_tvd_ft: np.ndarray,
) -> pd.DataFrame:
	ch0 = int(cfg['ch_shallow'])
	ch1 = int(cfg['ch_deep'])

	ch_min = min(ch0, ch1)
	ch_max = max(ch0, ch1)

	if keep_0based_incl is not None:
		k0, k1 = int(keep_0based_incl[0]), int(keep_0based_incl[1])
		if k1 < k0:
			raise ValueError(f'Bad keep_0based_incl: {keep_0based_incl}')
		if not (int(ch_min) <= int(k0) <= int(k1) <= int(ch_max)):
			raise ValueError(
				f'keep_0based_incl={keep_0based_incl} must be within endpoint channel range [{ch_min},{ch_max}]'
			)
		ch_min, ch_max = int(k0), int(k1)

	ch = np.arange(int(ch_min), int(ch_max) + 1, dtype=int)

	depth_m = channel_to_depth_m(ch, cfg)
	elev_m = float(cfg['elev_head_m']) - depth_m

	offE_m, offN_m = depth_to_offsets_m(depth_m, traj_ew_ft, traj_ns_ft, traj_tvd_ft)
	E_m = float(cfg['E_m']) + offE_m
	N_m = float(cfg['N_m']) + offN_m

	x_km = (E_m - float(origin_E_m)) / 1000.0
	y_km = (N_m - float(origin_N_m)) / 1000.0

	z_depth_km = depth_m / 1000.0
	z_elev_km = -elev_m / 1000.0

	index = int(index_offset) + (ch - int(ch_min))
	station_id = [f'{station_prefix}_CH{c:04d}' for c in ch.tolist()]

	# lon/lat（LOKI header用に便利）
	tr = Transformer.from_crs(f'EPSG:{int(INPUT_EPSG)}', 'EPSG:4326', always_xy=True)
	lon, lat = tr.transform(E_m.astype(float), N_m.astype(float))

	df = (
		pd.DataFrame(
			{
				'station_id': station_id,
				'well': cfg['well'],
				'channel': ch.astype(int),
				'index': index.astype(int),
				'E_m': E_m.astype(float),
				'N_m': N_m.astype(float),
				'lon': np.asarray(lon, dtype=float),
				'lat': np.asarray(lat, dtype=float),
				'elev_m': elev_m.astype(float),
				'depth_m': depth_m.astype(float),
				'x_km': x_km.astype(float),
				'y_km': y_km.astype(float),
				'z_depth_km': z_depth_km.astype(float),
				'z_elev_km': z_elev_km.astype(float),
			}
		)
		.sort_values('channel')
		.reset_index(drop=True)
	)

	return df


keep_a = tuple(WELL_A_KEEP_0BASED_INCL) if bool(APPLY_WELL_AB_KEEP) else None
keep_b = tuple(WELL_B_KEEP_0BASED_INCL) if bool(APPLY_WELL_AB_KEEP) else None

ewA_ft, nsA_ft, tvdA_ft = read_trajectory_offsets_ft(WELL_TRAJ_XLSX, SHEET_A)
ewB_ft, nsB_ft, tvdB_ft = read_trajectory_offsets_ft(WELL_TRAJ_XLSX, SHEET_B)

df_a = make_well_df(
	W78A,
	ORIGIN_E_m,
	ORIGIN_N_m,
	keep_0based_incl=keep_a,
	index_offset=0,
	station_prefix='DAS78A',
	traj_ew_ft=ewA_ft,
	traj_ns_ft=nsA_ft,
	traj_tvd_ft=tvdA_ft,
)

df_b = make_well_df(
	W78B,
	ORIGIN_E_m,
	ORIGIN_N_m,
	keep_0based_incl=keep_b,
	index_offset=int(df_a.shape[0]),
	station_prefix='DAS78B',
	traj_ew_ft=ewB_ft,
	traj_ns_ft=nsB_ft,
	traj_tvd_ft=tvdB_ft,
)

df = (
	pd.concat([df_a, df_b], ignore_index=True)
	.sort_values(['well', 'channel'])
	.reset_index(drop=True)
)
df.to_csv(OUT_CSV, index=False)

memo = f"""# FORGE DAS station metadata (78A-32 / 78B-32)

## What this is
Station table for GaMMA / LOKI.

- channel -> depth: tap-test calibrated linear map
- (E_m, N_m): **well trajectory offsets** (E-W, N-S vs TVD) from {WELL_TRAJ_XLSX.name} with interpolation
- lat/lon: converted from EPSG:{INPUT_EPSG} -> EPSG:4326

## Channel selection
APPLY_WELL_AB_KEEP={APPLY_WELL_AB_KEEP}
- 78A: {WELL_A_KEEP_0BASED_INCL}
- 78B: {WELL_B_KEEP_0BASED_INCL}

## Origin for local coordinates
E0={ORIGIN_E_m} m, N0={ORIGIN_N_m} m
x_km=(E-E0)/1000, y_km=(N-N0)/1000

## Columns
station_id, well, channel, index, E_m, N_m, lon, lat, elev_m, depth_m, x_km, y_km, z_depth_km, z_elev_km
"""
OUT_MEMO.write_text(memo, encoding='utf-8')

print(f'Wrote: {OUT_CSV.resolve()}')
print(f'Wrote: {OUT_MEMO.resolve()}')
print(df.head(2))
print(df[df['well'] == '78B-32'].tail(2))
# %%
